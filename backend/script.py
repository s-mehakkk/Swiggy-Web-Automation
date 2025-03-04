from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import locators as locators
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from selenium.webdriver.chrome.options import Options
import requests
import os
import subprocess

chrome_options = Options()
chrome_options.add_argument("--disable-background-timer-throttling")
chrome_options.add_argument("--disable-backgrounding-occluded-windows")
chrome_options.add_argument("--disable-renderer-backgrounding")
chrome_options.add_argument("--headless") 

URL = "https://www.swiggy.com/"

CHROME_PATH = "/opt/google/chrome/chrome"
CHROMEDRIVER_PATH = "/opt/chromedriver/chromedriver"

# Function to install Chrome
def install_chrome():
    print("Installing Google Chrome...")
    subprocess.run("mkdir -p /opt/google/chrome", shell=True, check=True)
    subprocess.run("wget -q https://dl.google.com/linux/direct/google-chrome-stable_current_amd64.deb", shell=True, check=True)
    subprocess.run("dpkg -x google-chrome-stable_current_amd64.deb /opt/google/", shell=True, check=True)

# Function to install ChromeDriver
def install_chromedriver():
    print("Installing ChromeDriver...")
    subprocess.run("mkdir -p /opt/chromedriver", shell=True, check=True)
    subprocess.run("wget -q https://chromedriver.storage.googleapis.com/114.0.5735.90/chromedriver_linux64.zip", shell=True, check=True)
    subprocess.run("unzip -o chromedriver_linux64.zip -d /opt/chromedriver/", shell=True, check=True)

# Install Chrome & ChromeDriver if not already installed
if not os.path.exists(CHROME_PATH):
    install_chrome()

if not os.path.exists(CHROMEDRIVER_PATH):
    install_chromedriver()

chrome_options.binary_location = CHROME_PATH

service = Service(CHROMEDRIVER_PATH)
driver = webdriver.Chrome(service=service, options=chrome_options)

print("Chrome & ChromeDriver successfully installed and running!")

first_outlet = True
Prev_location = ""

BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # This gets the swiggy-auto folder
API_DIR = os.path.join(BASE_DIR, "selenium-api")

def wait_for_input(filename):
    """Waits for user input from API"""
    file_path = os.path.join(os.getcwd(), filename)
    while not os.path.exists(file_path):
        print(f"Waiting for {filename}...")
        time.sleep(2)

    with open(file_path, "r") as file:
        value = file.read().strip()

    os.remove(file_path)
    return value

def open_and_login():
    driver.get(URL)

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, locators.Sign_in_span)))
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, locators.Sign_in_span))).click()

    # Wait for user to send phone number via API
    print("Waiting for user to send phone number...")
    phone_number = wait_for_input("phone.txt")
    print(f"Received phone number: {phone_number}")

    # Enter phone number and click login
    phone_input = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@type="tel"]')))
    phone_input.send_keys(phone_number)

    submit_button = driver.find_element(By.XPATH, '//a[contains(text(), "Login")]')
    submit_button.click()
    
    # Wait for OTP input from API
    print("Waiting for OTP input...")
    otp = wait_for_input("otp.txt")
    print(f"Received OTP: {otp}")

    # Enter OTP and submit
    otp_input = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@name="otp"]')))
    otp_input.send_keys(otp)

    otp_submit = driver.find_element(By.XPATH, '//a[contains(text(), "VERIFY OTP")]')
    otp_submit.click()

    print("Login successful. Proceeding with data extraction...")


def get_data_script(restaurants_data, OUTLET, OUTLET_NAME, detailDiscount=False):
    driver.get("https://www.swiggy.com/restaurants")
    global first_outlet
    if(first_outlet):
        try: 
            popup1 = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, locators.popup1_Xpath)))
            popup1.click()
        except TimeoutException:
            print("Popup1 not found, continuing...")
        try:
            popup2 = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, locators.popup2_Xpath)))
            popup2.click()
        except TimeoutException:
            print("Popup2 not found, continuing...")
        
        first_outlet = False

    restaurant_data = {}
    restaurant_data["Outlet Name"] = OUTLET_NAME
    restaurant_data["Outlet"] = OUTLET

    global Prev_location
    if(True or OUTLET != Prev_location):
        Location_element = driver.find_element(By.CLASS_NAME, locators.Location_class)
        Location_element.click()

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, locators.Location_input_Xpath)))
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, locators.Location_input_Xpath)))
        Location_input_element = driver.find_element(By.XPATH, locators.Location_input_Xpath)
        Location_input_element.send_keys(OUTLET)

        time.sleep(2) #TODO PUT SOME OTHER WAIT
        First_location_in_list = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, locators.First_location_in_list_Xpath)))
        First_location_in_list.click()
        try:
            Skip_and_add_later_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, locators.Skip_and_add_later_Xpath)))
            Skip_and_add_later_element.click()
        except:
            print("Skip and add later button not found, continuing...")
        Prev_location = OUTLET

    Search_element = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.XPATH, locators.Search_Xpath)))
    Search_element = WebDriverWait(driver, 6).until(EC.element_to_be_clickable((By.XPATH, locators.Search_Xpath)))
    Search_element.click()

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, locators.Search_input_Xpath)))
    Search_input_element = driver.find_element(By.XPATH, locators.Search_input_Xpath)
    Search_input_element.send_keys(OUTLET_NAME)

    First_restaurant_in_list = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, locators.First_restaurant_in_list_Xpath)))
    First_restaurant_in_list.click()

    Resultant_restaurant_element = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, locators.Resultant_restaurant_Xpath)))
    # checking if desired restaurant
    expected_name = OUTLET_NAME.strip().lower()
    Resultant_restaurant_name = driver.find_element(By.XPATH, locators.Resultant_restaurant_name_Xpath).text.strip().lower()
    if(Resultant_restaurant_name != expected_name):
        print("Restraunt name not matched")
        print(Resultant_restaurant_name)
        return

    try:
        closed_element = driver.find_element(By.XPATH, locators.Closed_resturant_Xpath)
        print(f"Skipping closed restaurant: {OUTLET_NAME} at {OUTLET}")
        processed_data = {
            "Name": restaurant_data["Outlet Name"],
            "Location": restaurant_data["Outlet"],
            "Rating": "",
            "CFT": "",
            "Tags": "",
            "Discounts": "",
            "Status": "Offline"
        }
        restaurants_data.append(processed_data)
        return
    
    except NoSuchElementException:
        pass
    
    Cuisine_tag_span = driver.find_element(By.XPATH, locators.Cuisine_tag_span_Xpath)
    restaurant_data["Tags"] = Cuisine_tag_span.text

    Resultant_restaurant_element.click()
    # Reached Restaurant Page
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, locators.Rating_div_Xpath)))
    Resultant_rating_element = driver.find_element(By.XPATH, locators.Rating_div_Xpath)
    restaurant_data["Rating"] = Resultant_rating_element.text

    Resultant_cft_element = driver.find_element(By.XPATH, locators.CFT_div_Xpath)
    restaurant_data["CFT"] = Resultant_cft_element.text
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, locators.Discount_div_Xpath)))
        Discount_divs = driver.find_elements(By.XPATH, locators.Discount_div_Xpath)
        discount_list = []
        for discount_div in Discount_divs:
            # 1. src attribute of the img in the first div
            img_div = discount_div.find_elements(By.TAG_NAME, "div")[0] 
            img_src = img_div.find_element(By.TAG_NAME, "img").get_attribute("src")

            if img_src == "https://media-assets.swiggy.com/swiggy/image/upload/fl_lossy,f_auto,q_auto,w_96,h_96/offers/generic":
                # 2. text of the nested divs in the second div
                if(detailDiscount):
                    time.sleep(1)
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(discount_div))
                    discount_div.click()
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, locators.Discount_detail_Xpath)))
                    text_div = driver.find_element(By.XPATH, locators.Discount_detail_Xpath)
                    text = text_div.text if len(text_div.text) > 0 else ""
                    discount_list.append([text])
                    driver.find_element(By.XPATH, locators.Discount_detail_close_Xpath).click()
                else:
                    text_div_container = discount_div.find_elements(By.TAG_NAME, "div")[1]
                    inner_divs = text_div_container.find_elements(By.TAG_NAME, "div")

                    text_1 = inner_divs[0].text if len(inner_divs) > 0 else ""
                    text_2 = inner_divs[1].text if len(inner_divs) > 1 else ""
                    # discount_list.append([text_1, text_2, img_src])
                    discount_list.append([text_1, text_2])
        restaurant_data["Discounts"] = discount_list
    except NoSuchElementException:
        pass

    discounts_str = "; ".join([f"{d[0]}: {d[1]}" if len(d) > 1 else f"{d[0]}" for d in restaurant_data["Discounts"]])
    processed_data = {
            "Name": restaurant_data["Outlet Name"],
            "Location": restaurant_data["Outlet"],
            "Rating": restaurant_data["Rating"],
            "CFT": restaurant_data["CFT"],
            "Tags": restaurant_data["Tags"],
            "Discounts": discounts_str,
            "Status": "Online"
    }
    restaurants_data.append(processed_data)
    

def save_excel(restaurants_data, restaurant):
    for entry in restaurants_data:
        entry["Discounts"] = entry["Discounts"].replace("; ", "\n")

    df = pd.DataFrame(restaurants_data)
    df.to_excel(f"./Results/{restaurant}_data.xlsx", index=False)

def modify_excel(restaurant):
    excel_file = f"./Results/{restaurant}_data.xlsx"
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)

        # Determine the maximum length of content in the column
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Set the column width
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_file)
    print(f"Adjusted Data saved to '{restaurant}_data.xlsx'")

def close_driver():
    driver.quit()

# Ensure Results directory exists
RESULTS_DIR = "./Results"
if not os.path.exists(RESULTS_DIR):
    os.makedirs(RESULTS_DIR)

print("Script started successfully on Render!")



# location, name
haus = [    
    ["Cyber City", "Asian Haus - By Haus Delivery"],
    ["Cyber City", "Sushi Haus - By Haus Delivery"],
    ["Greater Kailash 2", "Amma's Haus - by Asian Haus"],
    ["Greater Kailash 2", "Asian Haus - By Haus Delivery"],
    ["Greater Kailash 2", "Masala Haus - By Asian Haus"],
    ["Greater Kailash 2", "Sushi Haus - By Haus Delivery"],
    ["Vasant Kunj", "Asian Haus - By Haus Delivery"],
    ["Vasant Kunj", "Sushi Haus - By Haus Delivery"]
]
ambrosia = [
    ["DLF phase 4 layla's", "Layla's Shawarma & Middle Eastern Kitchen"],
    ["Green Park Layla's", "Layla's Shawarma & Middle Eastern Kitchen"],
    ["Greater Kailash I", "Layla's Shawarma & Middle Eastern Kitchen"],
    ["Punjabi Bagh layla's", "Layla's Shawarma & Middle Eastern Kitchen"]
]
how = [
    ["Vasant Kunj House of wok", "House of wok"],
    ["Greater Kailash I House of wok", "House of wok"],
    ["DLF phase 4 House of wok", "House of wok"],
    ["Baani Square", "House of wok"],
    ["Mohali House of wok", "House of wok"],
    ["Punjabi Bagh House of wok Prive", "House of wok Prive"],
    ["Iris Broadway", "House of wok"],
    ["M3m 65th avenue", "House of wok Prive"],
    ["Malviya nagar jaipur House of wok", "House of wok"]
]

pnb = [
    ["Vasant Kunj Punjabi by nature", "Punjabi by nature"],
    ["Sector 72 Noida", "Punjabi by nature"],
    ["DLF Phase 3 Punjabi by nature", "Punjabi by nature"],
    ["Hajipur Punjabi by nature", "Punjabi by nature"],
    ["Dehradun Punjabi by nature", "Punjabi by nature"],
    ["Punjabi by nature IFC", "Punjabi by nature"],
    ["DLF Phase 4", "Punjabi by nature"]
]

Hudson = [
    ["Hudson Chopsticks GTB nagar", "Hudson Chopsticks - Fresh Chinese"],
    ["Doner & Gyros GTB nagar", "Doner & Gyros - Salad, Shawarma, Falafel House"],
    ["Ashok Vihar", "Dragon Wok - Chinese restaurant"]
]

restaurants = [
    # {"name": "Ambrosia", "data": ambrosia, "detailDiscount": False},
    # {"name": "PunjabiByNature", "data": pnb, "detailDiscount": True},
    # {"name": "Haus", "data": haus, "detailDiscount": False},
    {"name": "Hudson", "data": Hudson, "detailDiscount": False},
    # {"name": "HouseOfWok", "data": how, "detailDiscount": False},
]

open_and_login()
# open_and_load_cookies()

for restaurant in restaurants:
    restaurants_data = []
    for outlet in restaurant["data"]:
        get_data_script(restaurants_data, outlet[0], outlet[1], restaurant["detailDiscount"])
    save_excel(restaurants_data, restaurant["name"])
    modify_excel(restaurant["name"])

close_driver() 